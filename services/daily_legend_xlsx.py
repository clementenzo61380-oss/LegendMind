"""Écriture / append d'un classeur Excel pour l'export quotidien Légende."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from services.legend_day_export import LegendDayStats

# One sheet per player (stable sheet name = tag without '#').
_HEADERS: tuple[str, ...] = (
    "Jour fin (UTC)",
    "Attaques",
    "Défenses",
    "ATK TR",
    "DEF TR",
    "+/- TR",
    "INIT TR",
    "FINAL TR",
)

_INDEX_HEADERS: tuple[str, ...] = (
    "Tag",
    "Nom",
    "Onglet",
)


def _sheet_name_for_tag(tag: str) -> str:
    return tag.strip().upper().lstrip("#")[:31] or "PLAYER"


def _ensure_index_sheet(wb: Workbook):
    if "INDEX" in wb.sheetnames:
        ws = wb["INDEX"]
        if ws.max_row < 1 or (ws["A1"].value is None):
            ws.append(list(_INDEX_HEADERS))
        return ws
    ws = wb.create_sheet(title="INDEX", index=0)
    ws.append(list(_INDEX_HEADERS))
    return ws


def _get_or_create_player_sheet(wb: Workbook, tag: str):
    name = _sheet_name_for_tag(tag)
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(list(_HEADERS))
    return name, ws


def append_legend_day_rows(path: Path, rows: list[LegendDayStats]) -> None:
    """Ajoute des stats quotidiennes dans un onglet par joueur (tag)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove default sheet.
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    index_ws = _ensure_index_sheet(wb)

    existing: dict[str, str] = {}
    for row in index_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        t = (row[0] or "").strip()
        s = (row[2] or "").strip() if len(row) >= 3 else ""
        if t and s:
            existing[t] = s

    for r in rows:
        sheet_name, ws = _get_or_create_player_sheet(wb, r.player_tag)
        ws.append(
            [
                r.legend_day_end_utc.isoformat(),
                r.attacks,
                r.defenses,
                r.trophies_attack,
                r.trophies_defense,
                r.net_trophies,
                r.trophies_start,
                r.trophies_end,
            ],
        )
        if r.player_tag not in existing:
            index_ws.append([r.player_tag, r.player_name or "", sheet_name])
            existing[r.player_tag] = sheet_name

    wb.save(path)
